
import scipy.io as sio
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
from tensorflow.keras.models import load_model
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Activation
from tensorflow.keras.optimizers import Adam
from tensorflow.keras import backend as K
import pickle
import sys
import os
import openpyxl
import random

random.seed(0)  # make the results reproducible

def r2_keras(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))

def training(X, y, X_test, y_test):
    # model structure
    # model = Sequential([
    #     Dense(32, input_dim=np.shape(X)[1]),
    #     Activation('relu'),
    #     Dense(16),
    #     Activation('relu'),
    #     Dense(1)
    # ])
    model = Sequential([
        Dense(128, input_dim=np.shape(X)[1]),
        Activation('relu'),
        Dense(32),
        Activation('relu'),
        Dense(1)
    ])

    # descent operator
    # adam = Adam(learning_rate=1e-3)
    adam = Adam(learning_rate=1e-3)
    # compile
    model.compile(optimizer=adam,
                  loss='mean_squared_error',  # loss function  mean_squared_error; mean_absolute_percentage_error
                  metrics=[r2_keras])  # metrics for test in training
    # training
    record = model.fit(X, y, epochs=100, batch_size=128, verbose=2,
                       validation_data=(X_test, y_test))
    return model, record


def draw_loss_function(record, saving=False, name=''):
    font = {'family': 'Times New Roman',
            'weight': 'normal',
            'size': 16}

    epochs = range(len(record.history['loss']))

    plt.figure()
    plt.plot(epochs, record.history['loss'], label='Training loss')
    plt.plot(epochs, record.history['val_loss'], label='Validation loss')
    plt.title('Training and validation loss', font)
    plt.legend()
    plt.show(block=False)
    if saving:
        plt.savefig(name + 'loss.png')

def draw_val_Y(y_test,y_test_pre, saving=False, name=''):
    # plot figure
    font = {'family': 'Times New Roman',
            'weight': 'normal',
            'size': 16}
    plt.figure()
    plt.scatter(np.array(y_test), np.array(y_test_pre), marker='o')
    plt.title('Validation of Y', font)
    plt.legend()
    plt.show(block=False)
    if saving:
        plt.savefig(name + '.png')


def NN_train(matfile, modelfile):
    # train a551 and Y separately
    matData = sio.loadmat(matfile)

    wavelength = matData['Syn']['wl'][0][0].squeeze()
    Rrs = pd.DataFrame(matData['Syn']['Rrs'][0][0], columns=wavelength)

    #  Reshape Data
    sat_band = [412, 443, 488, 547, 667]  # MODIS satellite bands
    # sat_band = [410, 443, 486, 530, 551, 620, 671]  # more satellite bands
    X_norm = Rrs.loc[:, sat_band]

    # output
    y_Y = pd.DataFrame(matData['Syn']['Y'][0][0], columns=['Y'])
    # y_Y = pd.DataFrame(matData['Y'], columns=['Y'])

    X_train, X_test, y_train, y_test = train_test_split(X_norm, y_Y, test_size=0.2, random_state=0)

    model_Y, record_Y = training(X_train, y_train, X_test, y_test)

    draw_loss_function(record_Y, saving=True, name='Y')

    # model save
    save_dir = './Model'
    os.makedirs(save_dir, exist_ok=True)
    # with open(save_dir + '/' + 'scaler_x_add2band.pkl', 'wb') as f:
    #     pickle.dump(scaler_x, f)
    model_Y.save(save_dir + '/' + modelfile)

    # model elevation
    y_test_pre = model_Y.predict(X_test)
    output_save = pd.DataFrame(np.concatenate([np.array(y_test),
                                               y_test_pre], axis=1))
    output_save.to_excel('Val_test_results0514.xlsx')
    # draw_val_Y(y_test, y_test_pre, saving=True, name='test_Y')


def NN_val(matfile, modelfile):
    # load model
    model_dir = './Model/'
    model = load_model(model_dir + modelfile, custom_objects={'r2_keras': r2_keras})

    # load synthetic data

    matData = sio.loadmat(matfile)
    wavelength = matData['Syn']['wl'][0][0].squeeze()
    Rrs = pd.DataFrame(matData['Syn']['Rrs'][0][0], columns=wavelength)
    y_Y = pd.DataFrame(matData['Syn']['Y'][0][0], columns=['Y'])
    sat_band = [412, 443, 488, 547, 667]  # MODIS satellite bands
    X_norm = Rrs.loc[:, sat_band]

    X_train, X_test, y_train, y_test = train_test_split(X_norm, y_Y, test_size=0.8, random_state=0)

    # Normalization
    y_train_predict = model.predict(X_train)
    # output
    output_save = pd.DataFrame(np.concatenate([np.array(y_train),
                                              y_train_predict], axis=1))
    output_save.to_excel('Val_resultsnewNNtoolddata0514.xlsx')

    # draw_val_Y(y_train, y_train_predict, saving=True, name='Validation_Y')


def NN_val_IOCCG(modelfile):

    # load field data
    matData = sio.loadmat('IOCCG.mat')
    wavelength = matData['field']['wl'][0][0].squeeze()
    Rrs = pd.DataFrame(matData['field']['Rrs'][0][0], columns=wavelength)
    y_Y = pd.DataFrame(matData['field']['Y'][0][0], columns=['Y'])

    # sat_band = [412, 443, 488, 547, 667]  # MODIS satellite bands

    LUT = np.array([[412, 0.003, 0.014, -0.022],
                    [443, 0.004, 0.015, -0.023],
                    [488, 0.011, 0.010, -0.051],
                    [551, 0.017, 0.010, -0.080],
                    [667, 0.018, 0.010, -0.081]])

    Rrs412 = Rrs.loc[:, 412]
    Rrs443 = Rrs.loc[:, 443]
    Rrs488 = Rrs.loc[:, 488]
    Rrs547 = Rrs.loc[:, 547]
    Rrs667 = Rrs.loc[:, 667]

    Rrs_ratio = Rrs443 / Rrs547
    iRrs412 = Rrs412 / (1 + (LUT[0][1] * Rrs_ratio + LUT[0][2] * Rrs547 ** LUT[0][3]))
    iRrs443 = Rrs443 / (1 + (LUT[1][1] * Rrs_ratio + LUT[1][2] * Rrs547 ** LUT[1][3]))
    iRrs488 = Rrs488 / (1 + (LUT[2][1] * Rrs_ratio + LUT[2][2] * Rrs547 ** LUT[2][3]))
    iRrs547 = Rrs547 / (1 + (LUT[3][1] * Rrs_ratio + LUT[3][2] * Rrs547 ** LUT[3][3]))
    iRrs667 = Rrs667 / (1 + (LUT[4][1] * Rrs_ratio + LUT[4][2] * Rrs547 ** LUT[4][3]))

    X_norm =pd.concat([iRrs412, iRrs443, iRrs488, iRrs547, iRrs667], axis=1)

    # load model
    model_dir = './Model/'
    model = load_model(model_dir + modelfile, custom_objects={'r2_keras': r2_keras})
    y_predict = model.predict(X_norm)

    output_save = pd.DataFrame(np.concatenate([np.array(y_Y),
                                              y_predict], axis=1))
    output_save.to_excel('Val_IOCCG0514.xlsx')

    # draw_val_Y(y_Y, y_predict, saving=True, name='IOCCG_Y')

def NN_val_field(modelfile):

    # load field data
    matData = sio.loadmat('Field.mat')
    wavelength = matData['field']['wl'][0][0].squeeze()
    Rrs = pd.DataFrame(matData['field']['Rrs'][0][0], columns=wavelength)
    y_Y = pd.DataFrame(matData['field']['Y'][0][0], columns=['Y'])

    # sat_band = [412, 443, 488, 547, 667]  # MODIS satellite bands

    LUT = np.array([[412, 0.003, 0.014, -0.022],
                    [443, 0.004, 0.015, -0.023],
                    [488, 0.011, 0.010, -0.051],
                    [551, 0.017, 0.010, -0.080],
                    [667, 0.018, 0.010, -0.081]])

    Rrs412 = Rrs.loc[:, 412]
    Rrs443 = Rrs.loc[:, 443]
    Rrs488 = Rrs.loc[:, 488]
    Rrs547 = Rrs.loc[:, 547]
    Rrs667 = Rrs.loc[:, 667]

    Rrs_ratio = Rrs443 / Rrs547
    iRrs412 = Rrs412 / (1 + (LUT[0][1] * Rrs_ratio + LUT[0][2] * Rrs547 ** LUT[0][3]))
    iRrs443 = Rrs443 / (1 + (LUT[1][1] * Rrs_ratio + LUT[1][2] * Rrs547 ** LUT[1][3]))
    iRrs488 = Rrs488 / (1 + (LUT[2][1] * Rrs_ratio + LUT[2][2] * Rrs547 ** LUT[2][3]))
    iRrs547 = Rrs547 / (1 + (LUT[3][1] * Rrs_ratio + LUT[3][2] * Rrs547 ** LUT[3][3]))
    iRrs667 = Rrs667 / (1 + (LUT[4][1] * Rrs_ratio + LUT[4][2] * Rrs547 ** LUT[4][3]))

    X_norm =pd.concat([iRrs412, iRrs443, iRrs488, iRrs547, iRrs667], axis=1)

    # load model
    model_dir = './Model/'
    model = load_model(model_dir + modelfile, custom_objects={'r2_keras': r2_keras})
    y_predict = model.predict(X_norm)
    output_save = pd.DataFrame(np.concatenate([np.array(y_Y),
                                               y_predict], axis=1))
    output_save.to_excel('Val_field0514.xlsx')

    # draw_val_Y(y_Y, y_predict, saving=True, name='Val_field_Y')

def save_excel(save_dir, save_name, value, title, sheet_name):
    """
    value: numpy value
    """
    ab_path = os.path.join(save_dir, save_name)
    os.makedirs(save_dir, exist_ok=True)
    if os.path.exists(ab_path):  # add new sheet
        wb = openpyxl.load_workbook(ab_path)
        sheet = wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name

    sheet.append(title)
    for row in range(value.shape[0]):
        for col in range(value.shape[1]):
            sheet.cell(row + 2, col + 1, value[row, col])
    wb.save(ab_path)


if __name__ == '__main__':
    matfile='SynData0511old.mat'
    modelfile='model_Y0514.h5'
    # NN_train(matfile, modelfile)
    NN_val(matfile, modelfile)
    # NN_val_IOCCG(modelfile)
    # NN_val_field(modelfile)
